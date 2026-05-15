import streamlit as st
import pandas as pd
import datetime
import plotly.express as px
import plotly.graph_objects as go
import sqlite3
import io
import os
import requests
import re
from werkzeug.security import generate_password_hash, check_password_hash

# ==========================================
# KONFIGURASI LOGO
# ==========================================
LOGO_PATH = "Logo.png"  # Ganti dengan nama file logo Anda jika berubah

# ==========================================
# CEK KETERSEDIAAN LIBRARY TAMBAHAN
# ==========================================
try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from wordcloud import WordCloud
    import matplotlib.pyplot as plt
    WORDCLOUD_AVAILABLE = True
except ImportError:
    WORDCLOUD_AVAILABLE = False

# ==========================================
# KONFIGURASI HALAMAN
# ==========================================
st.set_page_config(page_title="Aplikasi Pelaporan Blokade", page_icon="🚧", layout="wide", initial_sidebar_state="expanded")

# ==========================================
# DATABASE LOKAL (SQLITE) SETTING (V4)
# ==========================================
# PERBAIKAN STEP 1: Menggunakan @st.cache_resource agar koneksi database
# dikelola oleh Streamlit secara aman untuk lingkungan multi-user.
# Koneksi lama (global variable) berisiko menyebabkan crash atau data corrupt
# jika ada lebih dari satu user mengakses aplikasi secara bersamaan.
@st.cache_resource
def get_connection():
    """
    Membuat dan mengembalikan koneksi SQLite yang dikelola oleh Streamlit.
    Fungsi ini hanya dijalankan SEKALI, hasilnya di-cache dan di-reuse.
    """
    conn = sqlite3.connect('blokade_pro_v3.db', check_same_thread=False)
    return conn

conn = get_connection()

# ==========================================
# CATATAN PERBAIKAN (CHANGELOG)
# ==========================================
# v3 → v4 (Refactor & Security Improvements):
# [STEP 1] Koneksi DB: global variable → @st.cache_resource (thread-safe)
# [STEP 2] Manajemen User: data_editor langsung → form ganti password dengan hash
# [STEP 3] Upload File: tanpa validasi → cek ukuran, sanitasi nama, timestamp
# [STEP 4] Import Data: langsung append → validasi kolom wajib sebelum import
# [STEP 5] Koordinat: 8 desa → 14 desa (semua dari LOKASI_DATA)
# ==========================================

def init_db():
    c = conn.cursor()
    # Tabel Users (Menyimpan password yang di-hash)
    c.execute('''CREATE TABLE IF NOT EXISTS users
                 (username TEXT PRIMARY KEY, password TEXT, role TEXT, nama_lengkap TEXT)''')
    
    # Tabel Audit Logs
    c.execute('''CREATE TABLE IF NOT EXISTS audit_logs
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, timestamp TEXT, user TEXT, action TEXT, details TEXT)''')
    
    # Cek apakah admin sudah ada, jika belum buat dengan password default '123' yang di-hash
    c.execute("SELECT * FROM users WHERE username='admin'")
    if not c.fetchone():
        default_pw = generate_password_hash("123")
        c.execute("INSERT INTO users VALUES (?, ?, ?, ?)", ('admin', default_pw, 'Administrator', 'Super Admin'))
        c.execute("INSERT INTO users VALUES (?, ?, ?, ?)", ('spv', default_pw, 'Supervisor', 'Jhon Doe (SPV)'))
        c.execute("INSERT INTO users VALUES (?, ?, ?, ?)", ('mgr', default_pw, 'Manager', 'Bapak Manager'))
        conn.commit()

    # Tabel Laporan Blokade
    try:
        df_cek = pd.read_sql("SELECT * FROM laporan LIMIT 1", conn)
    except:
        DUMMY_DATA = [
            {"Tanggal": datetime.date(2024, 2, 10), "Pelaku": "Andi", "No HP": "0812345678", "Kabupaten": "Bolaang Mongondow", "Kecamatan": "Lolayan", "Desa": "Bakan", "Lokasi": "Simpang 3 Masjid Bakan", "Waktu Mulai": "08:00", "Waktu Selesai": "13:00", "Durasi (Jam)": 5.0, "Kategori Durasi": "Cepat", "Isu": "Tuntutan Pekerjaan", "Target": "PT JRBM & PT SMA", "Deskripsi": "Masyarakat menuntut pekerjaan untuk warga lokal. Tenda didirikan di simpang 3.", "File_Bukti": ""},
            {"Tanggal": datetime.date(2024, 5, 20), "Pelaku": "Budi", "No HP": "0812345678", "Kabupaten": "Bolaang Mongondow", "Kecamatan": "Lolayan", "Desa": "Matali Baru", "Lokasi": "Lokasi Umum Matali Baru", "Waktu Mulai": "07:00", "Waktu Selesai": "18:00", "Durasi (Jam)": 11.0, "Kategori Durasi": "Lambat", "Isu": "Kualitas Lingkungan", "Target": "PT JRBM", "Deskripsi": "Kompensasi debu jalan belum dibayarkan. Warga memblokade jalan pakai kayu.", "File_Bukti": ""},
            {"Tanggal": datetime.date(2024, 8, 15), "Pelaku": "Citra", "No HP": "0812345678", "Kabupaten": "Bolaang Mongondow Selatan", "Kecamatan": "Pinolosian Tengah", "Desa": "Tobayagan", "Lokasi": "Port Motandoi", "Waktu Mulai": "09:00", "Waktu Selesai": "13:30", "Durasi (Jam)": 4.5, "Kategori Durasi": "Cepat", "Isu": "Kompensasi Lahan", "Target": "PT JRBM", "Deskripsi": "Ganti rugi lahan area pelabuhan. Mediasi berhasil dilakukan jam 1 siang.", "File_Bukti": ""},
            {"Tanggal": datetime.date(2025, 1, 15), "Pelaku": "Eko", "No HP": "0812345678", "Kabupaten": "Bolaang Mongondow", "Kecamatan": "Lolayan", "Desa": "Bakan", "Lokasi": "Parkiran Blok C", "Waktu Mulai": "06:00", "Waktu Selesai": "10:00", "Durasi (Jam)": 4.0, "Kategori Durasi": "Cepat", "Isu": "Tuntutan Pekerjaan", "Target": "PT JRBM & PT SMA", "Deskripsi": "Protes penerimaan karyawan baru. Karyawan lokal merasa diabaikan.", "File_Bukti": ""},
            {"Tanggal": datetime.date(2025, 6, 20), "Pelaku": "Gita", "No HP": "0812345678", "Kabupaten": "Bolaang Mongondow Selatan", "Kecamatan": "Pinolosian Tengah", "Desa": "Tobayagan", "Lokasi": "Simpang 2 Akses Masuk Motandoi", "Waktu Mulai": "08:00", "Waktu Selesai": "20:00", "Durasi (Jam)": 12.0, "Kategori Durasi": "Lambat", "Isu": "Ganti Rugi Tanaman/Banjir", "Target": "PT JRBM", "Deskripsi": "Tuntutan ganti rugi tanaman cengkeh yang terkena dampak. Negosiasi alot.", "File_Bukti": ""},
        ]
        df_dummy = pd.DataFrame(DUMMY_DATA)
        df_dummy.to_sql('laporan', conn, index=False)
        add_audit_log("System", "INIT", "Sistem menginisialisasi database laporan pertama kali.")

def load_data():
    return pd.read_sql("SELECT * FROM laporan", conn)

def save_new_data(new_dict):
    new_df = pd.DataFrame([new_dict])
    new_df.to_sql('laporan', conn, if_exists='append', index=False)

def add_audit_log(user, action, details):
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c = conn.cursor()
    c.execute("INSERT INTO audit_logs (timestamp, user, action, details) VALUES (?, ?, ?, ?)", (now, user, action, details))
    conn.commit()

def send_telegram_notif(data_dict, pelapor):
    # Mengambil rahasia dari st.secrets jika tersedia (untuk cloud), jika tidak gunakan default/dummy
    try:
        BOT_TOKEN = st.secrets["BOT_TOKEN"]
        CHAT_ID = st.secrets["CHAT_ID"]
    except Exception:
        BOT_TOKEN = "DUMMY_TOKEN"
        CHAT_ID = "DUMMY_CHAT_ID"
    
    pesan = f"🚨 *LAPORAN BLOKADE BARU* 🚨\n\n"
    pesan += f"📍 *Lokasi:* {data_dict['Desa']}, {data_dict['Kabupaten']}\n"
    pesan += f"🏢 *Target:* {data_dict['Target']}\n"
    pesan += f"📢 *Isu/Tuntutan:* {data_dict['Isu']}\n"
    pesan += f"⏱️ *Durasi:* {data_dict['Durasi (Jam)']} Jam ({data_dict['Kategori Durasi']})\n"
    pesan += f"📝 *Deskripsi:* {data_dict['Deskripsi']}\n\n"
    pesan += f"👤 *Dilaporkan oleh:* {pelapor}"

    if BOT_TOKEN == "DUMMY_TOKEN":
        return False 

    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        payload = {"chat_id": CHAT_ID, "text": pesan, "parse_mode": "Markdown"}
        response = requests.post(url, json=payload)
        return response.status_code == 200
    except:
        return False

def authenticate(username, password):
    c = conn.cursor()
    c.execute("SELECT password, role, nama_lengkap FROM users WHERE username=?", (username,))
    user_record = c.fetchone()
    
    if user_record:
        hashed_password = user_record[0]
        # Memeriksa password menggunakan werkzeug security
        if check_password_hash(hashed_password, password):
            add_audit_log(user_record[2], "LOGIN", f"User {username} berhasil login.")
            return user_record[1], user_record[2] # Return Role dan Nama Lengkap
    return None

init_db()

# Buat folder upload jika belum ada
if not os.path.exists("uploads"):
    os.makedirs("uploads")

# ==========================================
# DATA MASTER (DICTIONARY)
# ==========================================
LOKASI_DATA = {
    "Bolaang Mongondow": {
        "Lolayan": {
            "Bakan": ["Simpang 3 Masjid Bakan", "Parkiran Blok C", "Simpang 3 Tapagale"],
            "Lolayan": ["Lokasi Umum Lolayan"],
            "Matali Baru": ["Lokasi Umum Matali Baru"],
            "Mopusi": ["Lokasi Umum Mopusi"]
        }
    },
    "Bolaang Mongondow Selatan": {
        "Pinolosian Tengah": {
            "Tobayagan": ["Port Motandoi", "Simpang 2 Akses Masuk Motandoi"],
            "Tobayagan Selatan": ["Lokasi Umum Tobayagan Selatan"]
        },
        "Pinolosian Timur": {
            "Motandoi": ["Lokasi Umum Motandoi"],
            "Motandoi Selatan": ["Lokasi Umum Motandoi Selatan"],
            "Dumagin A": ["Lokasi Umum Dumagin A"],
            "Dumagin B": ["Lokasi Umum Dumagin B"],
            "Onggunoi Induk": ["Lokasi Umum Onggunoi Induk"],
            "Onggunoi Selatan": ["Lokasi Umum Onggunoi Selatan"],
            "Pidung": ["Lokasi Umum Pidung"],
            "Dayow": ["Lokasi Umum Dayow"]
        }
    }
}

TUNTUTAN_DATA = {
    "Ganti Rugi Tanaman/Banjir": "PT JRBM",
    "Kompensasi Lahan": "PT JRBM",
    "Tuntutan Pekerjaan": "PT JRBM & PT SMA",
    "Kualitas Lingkungan": "PT JRBM",
    "Sanksi Indisipliner / Hubungan Industrial": "PT SMA",
    "Isu Sosial": "PT JRBM",
    "Lainnya": "Lainnya"
}

def get_coordinates(desa, kab):
    # ------------------------------------------------------------------
    # PERBAIKAN STEP 5: Melengkapi koordinat untuk semua desa yang ada
    # di dalam LOKASI_DATA. Kode lama hanya mendefinisikan 8 desa,
    # sehingga desa-desa lain jatuh ke koordinat fallback kabupaten dan
    # menumpuk di satu titik di peta (tidak akurat dan menyesatkan).
    # Koordinat di bawah adalah estimasi berdasarkan posisi geografis
    # aktual masing-masing desa di wilayah BM & BMS, Sulawesi Utara.
    # ------------------------------------------------------------------
    coords = {
        # === Kabupaten Bolaang Mongondow - Kecamatan Lolayan ===
        "Bakan"         : (0.6553, 124.0502),
        "Lolayan"       : (0.6700, 124.0100),
        "Matali Baru"   : (0.6300, 123.9800),
        "Mopusi"        : (0.6900, 124.0300),
        # === Kabupaten Bolaang Mongondow Selatan - Kec. Pinolosian Tengah ===
        "Tobayagan"         : (0.3800, 123.9500),
        "Tobayagan Selatan" : (0.3700, 123.9600),
        # === Kabupaten Bolaang Mongondow Selatan - Kec. Pinolosian Timur ===
        "Motandoi"          : (0.4000, 124.0000),
        "Motandoi Selatan"  : (0.3900, 123.9900),
        "Dumagin A"         : (0.3500, 123.9000),
        "Dumagin B"         : (0.3450, 123.8950),
        "Onggunoi Induk"    : (0.3600, 123.9100),
        "Onggunoi Selatan"  : (0.3550, 123.9050),
        "Pidung"            : (0.3650, 123.9200),
        "Dayow"             : (0.3750, 123.9300),
    }
    if desa in coords:
        return coords[desa]
    # Fallback: jika desa belum terdaftar, gunakan pusat kabupaten
    if kab == "Bolaang Mongondow":
        return (0.6500, 124.0000)
    return (0.3800, 123.9500)  # Fallback BMS

# ==========================================
# INISIALISASI STATE (LOGIN)
# ==========================================
if 'logged_in' not in st.session_state: st.session_state['logged_in'] = False
if 'role' not in st.session_state: st.session_state['role'] = None
if 'nama_lengkap' not in st.session_state: st.session_state['nama_lengkap'] = None

# ==========================================
# CSS CUSTOM TERMASUK @MEDIA PRINT
# ==========================================
def inject_custom_css():
    st.markdown("""
        <style>
        .stApp, .main .block-container { background-color: #f8f9fa !important; }
        section[data-testid="stSidebar"] { background-color: #ffffff !important; border-right: 1px solid #e2e8f0 !important; }
        p, h1, h2, h3, h4, h5, h6, span, label, div.stMarkdown, .stSelectbox label, .stTextInput label, .stDateInput label, .stTimeInput label { color: #000000 !important; }

        .metric-card {
            padding: 20px; border-radius: 12px; margin-bottom: 20px;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1); font-family: 'Segoe UI', Tahoma;
            position: relative; overflow: hidden; border: 1px solid rgba(255,255,255,0.2);
        }
        .metric-card p, .metric-card div, .metric-card span { color: #ffffff !important; text-shadow: 1px 1px 2px rgba(0,0,0,0.2); }
        
        /* Palet Warna KPI yang dikembalikan */
        .card-bm    { background: linear-gradient(135deg, #3b82f6, #2563eb); } 
        .card-bms   { background: linear-gradient(135deg, #ec4899, #db2777); } 
        .card-sma   { background: linear-gradient(135deg, #f59e0b, #d97706); } 
        .card-jrbm  { background: linear-gradient(135deg, #10b981, #059669); } 
        .card-lost  { background: linear-gradient(135deg, #ef4444, #dc2828); } 
        .card-total { background: linear-gradient(135deg, #6366f1, #4f46e5); } 
        .card-cepat { background: linear-gradient(135deg, #06b6d4, #0891b2); } 
        .card-lambat{ background: linear-gradient(135deg, #f97316, #ea580c); } 
        .card-avg-bm { background: linear-gradient(135deg, #8b5cf6, #7c3aed); } 
        .card-avg-bms { background: linear-gradient(135deg, #64748b, #475569); } 
        
        .metric-title { font-size: 13px; opacity: 0.95; margin-bottom: 8px; font-weight: 600; text-transform: uppercase; }
        .metric-value { font-size: 34px; font-weight: 800; line-height: 1.2;}
        .metric-subtitle { font-size: 12px; opacity: 0.85; margin-top: 6px; }

        button span, button p, .stAlert p, .stAlert span { color: inherit !important; }

        .login-box {
            background-color: #ffffff; padding: 40px; border-radius: 16px;
            box-shadow: 0 10px 25px rgba(0,0,0,0.05); border: 1px solid #e2e8f0; margin-top: 20px;
        }
        .sidebar-profile {
            background: linear-gradient(135deg, #0f172a, #1e293b); padding: 20px;
            border-radius: 12px; text-align: center; margin-bottom: 25px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }
        .sidebar-profile p, .sidebar-profile h4, .sidebar-profile div { color: #ffffff !important; }
        
        .summary-box {
            background-color: #e0f2fe; border-left: 5px solid #0284c7; padding: 15px 20px;
            border-radius: 8px; margin-bottom: 20px; color: #0c4a6e !important;
        }

        @media print {
            section[data-testid="stSidebar"] { display: none !important; } 
            header[data-testid="stHeader"] { display: none !important; }   
            .stButton, button, .stDownloadButton { display: none !important; } 
            .main .block-container { max-width: 100% !important; padding: 0 !important; margin: 0 !important; }
            * { -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }
            .js-plotly-plot, .metric-card { page-break-inside: avoid; margin-bottom: 15px;}
        }
        </style>
    """, unsafe_allow_html=True)

def force_black_text_on_plot(fig):
    fig.update_layout(
        paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
        font=dict(color="#000000", family="Segoe UI"),
        title_font=dict(color="#000000", size=16, family="Segoe UI", weight="bold"),
        legend_font=dict(color="#000000"),
        margin=dict(l=65, r=20, t=50, b=50)
    )
    fig.update_xaxes(showgrid=False, title_font=dict(color="#000000"), tickfont=dict(color="#000000"))
    fig.update_yaxes(showgrid=True, gridcolor="#e2e8f0", title_font=dict(color="#000000"), tickfont=dict(color="#000000"), title_standoff=15)
    return fig

# ==========================================
# FUNGSI MEMBUAT KARTU KPI
# ==========================================
def create_kpi_card(title, value, subtitle, css_class):
    return f"""
    <div class="metric-card {css_class}">
        <div class="metric-title">{title}</div>
        <div class="metric-value">{value}</div>
        <div class="metric-subtitle">{subtitle}</div>
    </div>
    """

# ==========================================
# FUNGSI EXPORT PDF & EXCEL
# ==========================================
def generate_pdf(df):
    pdf = FPDF(orientation='L', unit='mm', format='A3')
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 10, "Laporan Detail Kejadian Blokade", ln=True, align='C')
    pdf.ln(5)
    
    pdf.set_font("Arial", 'B', 9)
    col_widths = [22, 25, 25, 35, 35, 30, 45, 20, 20, 20, 25, 45]
    columns = list(df.columns)
    
    for i, col_name in enumerate(columns):
        w = col_widths[i] if i < len(col_widths) else 30
        pdf.cell(w, 8, str(col_name), border=1, align='C')
    pdf.ln()

    pdf.set_font("Arial", '', 8)
    for _, row in df.iterrows():
        for i, value in enumerate(row):
            w = col_widths[i] if i < len(col_widths) else 30
            text_val = str(value)[:50] + ".." if len(str(value)) > 50 else str(value)
            pdf.cell(w, 8, text_val, border=1, align='C')
        pdf.ln()
    return pdf.output(dest='S').encode('latin-1')

def generate_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Laporan Blokade')
    return output.getvalue()

# ==========================================
# FUNGSI AUTO-SUMMARY (CERDAS)
# ==========================================
def generate_smart_summary(df):
    if df.empty:
        return "Tidak ada data pada periode ini untuk dianalisis."
    
    total_kasus = len(df)
    total_jam = df['Durasi (Jam)'].sum()
    top_isu = df['Isu'].mode()[0] if not df['Isu'].empty else "-"
    top_kabupaten = df['Kabupaten'].mode()[0] if not df['Kabupaten'].empty else "-"
    top_desa = df['Desa'].mode()[0] if not df['Desa'].empty else "-"
    lambat_pct = (len(df[df['Kategori Durasi'] == 'Lambat']) / total_kasus) * 100 if total_kasus > 0 else 0

    narasi = f"Pada periode yang difilter, tercatat sebanyak **{total_kasus} insiden blokade** yang mengakibatkan terbuangnya waktu operasional (*Lost Time*) sebesar **{total_jam:.1f} jam**. "
    narasi += f"Secara demografis, area paling rawan saat ini berpusat di wilayah **{top_kabupaten} (khususnya Desa {top_desa})**. "
    narasi += f"Akar masalah utama yang paling sering memicu insiden adalah terkait **'{top_isu}'**. "
    
    if lambat_pct > 50:
        narasi += f"⚠️ **Perhatian:** Sekitar **{lambat_pct:.0f}%** dari total insiden masuk dalam kategori penanganan lambat (> 9 Jam), mengindikasikan perlunya evaluasi strategi respons tim di lapangan."
    else:
        narasi += f"✅ Mayoritas insiden berhasil ditangani dengan respons cepat (≤ 9 Jam), menunjukkan performa tim mediasi lapangan yang cukup efektif."
        
    return narasi

# ==========================================
# FUNGSI-FUNGSI UI (HALAMAN-HALAMAN)
# ==========================================

def login_page():
    inject_custom_css()
    st.markdown("<br><br>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        # Menampilkan Logo dengan ukuran proporsional di form login
        try:
            st.image(LOGO_PATH, use_container_width=True)
        except Exception:
            pass # Lewati jika file logo belum ada/salah nama

        st.markdown("""
        <div class="login-box">
            <h2 style='text-align: center; font-weight: 800; color: #1e293b !important;'>PORTAL BLOKADE</h2>
            <p style='text-align: center; color: #64748b !important; margin-bottom: 25px;'>Sistem Pelaporan & Analitik Interaktif</p>
        """, unsafe_allow_html=True)
        
        with st.form("login_form"):
            username = st.text_input("👤 Username")
            password = st.text_input("🔑 Password", type="password")
            submit = st.form_submit_button("🚀 Masuk Ke Sistem", use_container_width=True)
            
            if submit:
                user_data = authenticate(username, password)
                if user_data:
                    st.session_state['logged_in'] = True
                    st.session_state['role'] = user_data[0] # Role
                    st.session_state['nama_lengkap'] = user_data[1] # Nama
                    st.rerun()
                else:
                    st.error("❌ Username atau Password salah!")
        
        # Keterangan kredensial telah dihapus dari sini sesuai permintaan
        st.markdown("</div>", unsafe_allow_html=True)

def input_form_page():
    # Logo telah dihapus dari sini dan dipindahkan ke st.sidebar di main()
    st.header("📝 Form Input Pelaporan Blokade")
    st.markdown("Silakan isi form di bawah ini dengan lengkap untuk mencatat insiden baru.")
    
    with st.container():
        st.subheader("🧑‍🤝‍🧑 Data Pelaku & Lokasi")
        col1, col2 = st.columns(2)
        with col1:
            nama_pelaku = st.text_input("👤 Nama Pelaku")
            no_hp = st.text_input("📱 Nomor HP Pelaku", help="Hanya masukkan angka")
        
        with col2:
            kabupaten = st.selectbox("📍 Kabupaten", list(LOKASI_DATA.keys()))
            kecamatan = st.selectbox("🏘️ Kecamatan", list(LOKASI_DATA[kabupaten].keys()))
            desa = st.selectbox("🏡 Desa", list(LOKASI_DATA[kabupaten][kecamatan].keys()))
            lokasi = st.selectbox("📌 Lokasi Kejadian", LOKASI_DATA[kabupaten][kecamatan][desa])

    st.markdown("---")
    
    with st.container():
        st.subheader("⏱️ Waktu Kejadian")
        col3, col4, col5 = st.columns(3)
        with col3:
            tanggal = st.date_input("📅 Tanggal Kejadian")
        with col4:
            waktu_mulai = st.time_input("⏳ Waktu Mulai")
        with col5:
            waktu_selesai = st.time_input("⌛ Waktu Selesai")
            
        dt_mulai = datetime.datetime.combine(tanggal, waktu_mulai)
        dt_selesai = datetime.datetime.combine(tanggal, waktu_selesai)
        if dt_selesai < dt_mulai: dt_selesai += datetime.timedelta(days=1)
            
        durasi_detik = (dt_selesai - dt_mulai).total_seconds()
        durasi_jam = round(durasi_detik / 3600, 2)
        kategori_durasi = "Cepat" if durasi_jam <= 9 else "Lambat"
        
        st.info(f"⚙️ **Kalkulasi Durasi Otomatis:** {durasi_jam} Jam (Kategori: **{kategori_durasi}**)")

    st.markdown("---")
    
    with st.container():
        st.subheader("🎯 Tuntutan & Follow Up")
        col6, col7 = st.columns(2)
        with col6:
            isu = st.selectbox("📢 Jenis Isu / Tuntutan", list(TUNTUTAN_DATA.keys()))
        with col7:
            target_perusahaan = st.text_input("🏢 Target Perusahaan (Otomatis)", value=TUNTUTAN_DATA[isu], disabled=True)
        
        deskripsi = st.text_area("✍️ Deskripsi / Follow Up Kejadian", height=100)
        
        bukti = st.file_uploader("📎 Upload Bukti Foto (Opsional)", type=['jpg', 'png', 'jpeg'])
        
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("💾 Simpan Data Blokade ke Database", type="primary", use_container_width=True):
            # Validasi form tambahan
            if nama_pelaku.strip() == "":
                st.error("❌ Gagal Menyimpan: Nama Pelaku harus diisi!")
            elif not re.match(r"^[0-9]+$", no_hp) and no_hp != "":
                st.error("❌ Gagal Menyimpan: Nomor HP hanya boleh berisi angka!")
            else:
                # ------------------------------------------------------------------
                # PERBAIKAN STEP 3: Validasi file upload yang lebih aman.
                # Kode lama langsung menyimpan file tanpa pengecekan apapun,
                # yang berisiko:
                # 1. File terlalu besar (tidak ada batas ukuran)
                # 2. Nama file berbahaya (path traversal attack)
                # 3. Nama file bertabrakan jika ada file dengan nama yang sama
                # Solusi: cek ukuran, sanitasi nama, dan tambahkan timestamp unik.
                # ------------------------------------------------------------------
                file_path = ""
                MAKS_UKURAN_MB = 5  # Batas maksimum upload: 5 MB
                TIPE_DIIZINKAN = ['jpg', 'jpeg', 'png']

                if bukti:
                    # 1. Cek ukuran file
                    ukuran_mb = len(bukti.getbuffer()) / (1024 * 1024)
                    ekstensi   = bukti.name.rsplit('.', 1)[-1].lower()

                    if ukuran_mb > MAKS_UKURAN_MB:
                        st.error(f"❌ File terlalu besar ({ukuran_mb:.1f} MB). Maksimum {MAKS_UKURAN_MB} MB.")
                        st.stop()
                    elif ekstensi not in TIPE_DIIZINKAN:
                        st.error(f"❌ Tipe file tidak diizinkan. Hanya: {', '.join(TIPE_DIIZINKAN)}")
                        st.stop()
                    else:
                        # 2. Sanitasi nama file: hapus karakter berbahaya
                        nama_aman = re.sub(r'[^a-zA-Z0-9_.-]', '_', bukti.name)
                        # 3. Tambahkan timestamp agar nama unik, hindari tabrakan
                        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                        nama_final = f"{timestamp}_{nama_aman}"
                        file_path  = os.path.join("uploads", nama_final)

                        with open(file_path, "wb") as f:
                            f.write(bukti.getbuffer())

                new_data = {
                    "Tanggal": str(tanggal), "Pelaku": nama_pelaku, "No HP": no_hp,
                    "Kabupaten": kabupaten, "Kecamatan": kecamatan, "Desa": desa, "Lokasi": lokasi,
                    "Waktu Mulai": str(waktu_mulai), "Waktu Selesai": str(waktu_selesai), "Durasi (Jam)": durasi_jam,
                    "Kategori Durasi": kategori_durasi, "Isu": isu, "Target": target_perusahaan, "Deskripsi": deskripsi,
                    "File_Bukti": file_path
                }
                save_new_data(new_data)
                add_audit_log(st.session_state['nama_lengkap'], "CREATE", f"Menambahkan laporan blokade baru di {desa}.")
                
                st.success("✅ Data Blokade Berhasil Disimpan Permanen ke Sistem!")
                
                st.toast("Mempersiapkan Notifikasi...", icon="⏳")
                is_sent = send_telegram_notif(new_data, st.session_state['nama_lengkap'])
                
                if is_sent: st.info("📲 Notifikasi Telegram berhasil dikirim ke Grup Manajemen!")
                else: st.info(f"🔔 **Simulasi Notifikasi:** Pesan peringatan terkirim ke Manajemen bahwa terjadi blokade di {desa}.")
                
                st.balloons()


def kelola_data_page():
    st.header("⚙️ Kelola Data & Pengaturan")
    
    tab1, tab2, tab3, tab4 = st.tabs(["📝 Edit & Hapus Laporan", "📥 Import Data Massal", "👥 Manajemen User", "🕵️‍♂️ Log Aktivitas"])
    
    with tab1:
        st.markdown("### Edit atau Hapus Data Laporan (Full CRUD)")
        st.caption("Ubah data langsung di dalam tabel di bawah ini, lalu klik tombol Simpan Perubahan di bagian bawah tabel.")
        df = load_data()
        
        if not df.empty:
            edited_df = st.data_editor(df, num_rows="dynamic", use_container_width=True)
            
            if st.button("💾 Simpan Perubahan Tabel Laporan", type="primary"):
                edited_df.to_sql('laporan', conn, if_exists='replace', index=False)
                add_audit_log(st.session_state['nama_lengkap'], "UPDATE/DELETE", "Mengubah atau menghapus baris data laporan.")
                st.success("Tabel Laporan berhasil diperbarui!")
                st.rerun()
        else:
            st.info("Belum ada data laporan.")

    with tab2:
        st.markdown("### 📥 Import Data Historis")
        st.caption("Unggah file CSV atau Excel yang berisi data laporan. Pastikan nama kolom sesuai standar aplikasi.")

        # ------------------------------------------------------------------
        # PERBAIKAN STEP 4: Validasi kolom sebelum data diimpor ke database.
        # Kode lama langsung mengeksekusi import tanpa memeriksa apakah kolom
        # di file yang diunggah sesuai dengan skema tabel database.
        # Akibatnya, data bisa masuk dengan format yang salah atau error
        # tanpa pesan yang jelas bagi pengguna.
        # Solusi: definisikan kolom wajib, bandingkan dengan file yang diupload,
        # dan tampilkan pesan error yang informatif sebelum import dieksekusi.
        # ------------------------------------------------------------------
        KOLOM_WAJIB = [
            "Tanggal", "Pelaku", "No HP", "Kabupaten", "Kecamatan", "Desa",
            "Lokasi", "Waktu Mulai", "Waktu Selesai", "Durasi (Jam)",
            "Kategori Durasi", "Isu", "Target", "Deskripsi", "File_Bukti"
        ]

        # Tampilkan panduan kolom yang diperlukan
        with st.expander("📋 Lihat Format Kolom yang Diperlukan"):
            st.code(", ".join(KOLOM_WAJIB))
            st.caption("Pastikan file Anda memiliki semua kolom di atas dengan nama yang sama persis (case-sensitive).")

        uploaded_file = st.file_uploader("Pilih file data (.csv atau .xlsx)", type=["csv", "xlsx"])
        
        if uploaded_file is not None:
            try:
                if uploaded_file.name.endswith('.csv'):
                    df_import = pd.read_csv(uploaded_file)
                else:
                    df_import = pd.read_excel(uploaded_file)
                
                # --- Validasi kolom ---
                kolom_ada      = set(df_import.columns.tolist())
                kolom_wajib    = set(KOLOM_WAJIB)
                kolom_kurang   = kolom_wajib - kolom_ada
                kolom_ekstra   = kolom_ada - kolom_wajib

                if kolom_kurang:
                    # Ada kolom wajib yang tidak ditemukan di file
                    st.error(f"❌ Import dibatalkan! Kolom berikut **tidak ditemukan** di file Anda:")
                    st.code(", ".join(sorted(kolom_kurang)))
                    st.info("💡 Silakan perbaiki file Anda dan coba lagi.")
                else:
                    # Semua kolom wajib ada
                    if kolom_ekstra:
                        st.warning(f"⚠️ Kolom tambahan di luar standar ditemukan dan akan **diabaikan**: `{', '.join(sorted(kolom_ekstra))}`")

                    st.success(f"✅ Format kolom valid! Ditemukan **{len(df_import)} baris** data siap diimpor.")
                    st.write("**Preview 5 baris pertama:**")
                    st.dataframe(df_import[KOLOM_WAJIB].head(), use_container_width=True)
                    
                    if st.button("🚀 Eksekusi Import Data", type="primary"):
                        # Hanya simpan kolom yang sesuai standar
                        df_import[KOLOM_WAJIB].to_sql('laporan', conn, if_exists='append', index=False)
                        add_audit_log(st.session_state['nama_lengkap'], "IMPORT",
                                      f"Mengimpor {len(df_import)} baris data dari file {uploaded_file.name}.")
                        st.success(f"✅ Berhasil mengimpor {len(df_import)} baris data!")
                        st.rerun()

            except Exception as e:
                st.error(f"❌ Terjadi kesalahan saat membaca file: {e}")

    with tab3:
        if st.session_state['role'] == 'Administrator':
            st.markdown("### 👥 Manajemen Akun Pengguna")

            # ------------------------------------------------------------------
            # PERBAIKAN STEP 2: Memisahkan fitur manajemen user menjadi dua
            # bagian yang aman:
            # 1. Tabel read-only untuk melihat daftar user (tanpa kolom password)
            # 2. Form khusus untuk ganti password dengan hash otomatis
            # Sebelumnya, user bisa langsung mengedit kolom password di tabel
            # yang bisa menyebabkan password tersimpan sebagai plaintext (tidak
            # ter-hash), sehingga login akan selalu gagal.
            # ------------------------------------------------------------------

            # --- Bagian 1: Tampilan daftar user (tanpa kolom password) ---
            st.markdown("#### 📋 Daftar Akun Terdaftar")
            st.caption("Kolom password disembunyikan untuk keamanan. Gunakan form di bawah untuk mengubah password.")
            df_users = pd.read_sql("SELECT username, role, nama_lengkap FROM users", conn)
            st.dataframe(df_users, use_container_width=True, hide_index=True)

            st.markdown("---")

            # --- Bagian 2: Form tambah user baru ---
            st.markdown("#### ➕ Tambah Akun Baru")
            with st.form("form_tambah_user", clear_on_submit=True):
                col_u1, col_u2 = st.columns(2)
                with col_u1:
                    new_username   = st.text_input("👤 Username Baru")
                    new_nama       = st.text_input("📛 Nama Lengkap")
                with col_u2:
                    new_role       = st.selectbox("🎭 Role", ["Administrator", "Supervisor", "Manager"])
                    new_pw         = st.text_input("🔑 Password Baru", type="password")
                    new_pw_confirm = st.text_input("🔑 Konfirmasi Password", type="password")

                submit_tambah = st.form_submit_button("➕ Tambah User", use_container_width=True)
                if submit_tambah:
                    if not new_username.strip() or not new_pw.strip():
                        st.error("❌ Username dan Password tidak boleh kosong!")
                    elif new_pw != new_pw_confirm:
                        st.error("❌ Konfirmasi password tidak cocok!")
                    else:
                        c = conn.cursor()
                        c.execute("SELECT username FROM users WHERE username=?", (new_username,))
                        if c.fetchone():
                            st.error(f"❌ Username '{new_username}' sudah ada!")
                        else:
                            hashed = generate_password_hash(new_pw)
                            c.execute("INSERT INTO users VALUES (?, ?, ?, ?)",
                                      (new_username, hashed, new_role, new_nama))
                            conn.commit()
                            add_audit_log(st.session_state['nama_lengkap'], "USER_CREATE",
                                          f"Menambahkan akun baru: {new_username} ({new_role}).")
                            st.success(f"✅ Akun '{new_username}' berhasil ditambahkan!")
                            st.rerun()

            st.markdown("---")

            # --- Bagian 3: Form ganti password ---
            st.markdown("#### 🔑 Ganti Password Akun")
            st.caption("Password akan di-hash secara otomatis sebelum disimpan, sehingga aman.")
            with st.form("form_ganti_password", clear_on_submit=True):
                col_p1, col_p2 = st.columns(2)
                with col_p1:
                    target_user  = st.selectbox("👤 Pilih Akun", df_users['username'].tolist())
                    new_password = st.text_input("🔑 Password Baru", type="password")
                with col_p2:
                    st.markdown("<br>", unsafe_allow_html=True)
                    confirm_password = st.text_input("🔑 Konfirmasi Password Baru", type="password")

                submit_pw = st.form_submit_button("💾 Simpan Password Baru", type="primary", use_container_width=True)
                if submit_pw:
                    if not new_password.strip():
                        st.error("❌ Password tidak boleh kosong!")
                    elif new_password != confirm_password:
                        st.error("❌ Konfirmasi password tidak cocok! Pastikan kedua input sama.")
                    elif len(new_password) < 6:
                        st.error("❌ Password minimal 6 karakter!")
                    else:
                        # Hash password sebelum disimpan - INILAH INTI PERBAIKANNYA
                        hashed_pw = generate_password_hash(new_password)
                        c = conn.cursor()
                        c.execute("UPDATE users SET password=? WHERE username=?", (hashed_pw, target_user))
                        conn.commit()
                        add_audit_log(st.session_state['nama_lengkap'], "USER_MGMT",
                                      f"Mengganti password untuk akun: {target_user}.")
                        st.success(f"✅ Password untuk akun '{target_user}' berhasil diperbarui!")
        else:
            st.error("🚫 Akses Ditolak: Anda harus login sebagai Administrator untuk melihat menu ini.")
            
    with tab4:
        if st.session_state['role'] in ['Administrator', 'Manager']:
            st.markdown("### 🕵️‍♂️ Log Riwayat Aktivitas (Audit Trail)")
            st.caption("Memantau setiap pergerakan data di dalam sistem untuk transparansi operasional.")
            df_logs = pd.read_sql("SELECT * FROM audit_logs ORDER BY id DESC", conn)
            st.dataframe(df_logs, use_container_width=True, hide_index=True)
        else:
            st.error("🚫 Akses Ditolak: Hanya Manager dan Administrator yang dapat melihat log audit.")


def dashboard_page():
    inject_custom_css()
    
    col_head1, col_head2 = st.columns([3, 1])
    with col_head1:
        st.header("📊 Dashboard Analitik Blokade")
        st.caption("Ringkasan Performa dan Analisa Kendala di Lapangan secara Real-time")
    with col_head2:
        st.markdown("""
            <div style="margin-top: 15px;">
                <button onclick="window.print()" style="background-color: #0f172a; color: white; padding: 12px 20px; border: none; border-radius: 8px; cursor: pointer; font-weight: bold; width: 100%; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    🖨️ Cetak Full Dashboard
                </button>
            </div>
        """, unsafe_allow_html=True)
    
    df = load_data()
    
    if df.empty:
        st.warning("⚠️ Belum ada data laporan di Database.")
        return

    df['Tanggal'] = pd.to_datetime(df['Tanggal']).dt.date
    df['Tahun'] = pd.to_datetime(df['Tanggal']).dt.year
    df['Bulan_Tahun'] = pd.to_datetime(df['Tanggal']).dt.strftime('%b %Y')
    df['Quarter'] = 'Q' + pd.to_datetime(df['Tanggal']).dt.quarter.astype(str)

    # FILTER SECTION
    st.markdown("#### 🔍 Filter Analitik Canggih")
    col_f1, col_f2, col_f3 = st.columns([1, 1, 1.5])
    with col_f1:
        selected_year = st.selectbox("📆 Pilih Tahun", ["Semua Tahun"] + list(sorted(df['Tahun'].unique(), reverse=True)))
    with col_f2:
        selected_q = st.selectbox("🗂️ Periode Triwulan", ["Semua", "Q1", "Q2", "Q3", "Q4"])
    with col_f3:
        date_range = st.date_input("🗓️ Filter Rentang Waktu (Awal - Akhir)", [])
    
    df_raw = df.copy()

    if selected_year != "Semua Tahun": df = df[df['Tahun'] == selected_year]
    if selected_q != "Semua": df = df[df['Quarter'] == selected_q]
    if len(date_range) == 2:
        start_date, end_date = date_range
        df = df[(df['Tanggal'] >= start_date) & (df['Tanggal'] <= end_date)]

    st.markdown("<hr/>", unsafe_allow_html=True)

    # RINGKASAN CERDAS (SMART SUMMARY)
    st.markdown("#### 🤖 Ringkasan Eksekutif Otomatis")
    smart_text = generate_smart_summary(df)
    st.markdown(f"<div class='summary-box'>{smart_text}</div>", unsafe_allow_html=True)

    # TREN YOY
    if selected_year != "Semua Tahun":
        df_prev = df_raw[df_raw['Tahun'] == (selected_year - 1)]
        tot_blokade_now = len(df); tot_blokade_prev = len(df_prev); delta_blokade = tot_blokade_now - tot_blokade_prev
        tot_jam_now = df['Durasi (Jam)'].sum(); tot_jam_prev = df_prev['Durasi (Jam)'].sum(); delta_jam = tot_jam_now - tot_jam_prev

        m1, m2, m3 = st.columns(3)
        m1.metric("🛡️ Total Blokade", f"{tot_blokade_now} Kasus", f"{delta_blokade} Kasus vs Tahun Lalu", "inverse")
        m2.metric("⏳ Total Lost Time", f"{tot_jam_now:.1f} Jam", f"{delta_jam:.1f} Jam vs Tahun Lalu", "inverse")
        m3.metric("⚡ Rata-rata Durasi", f"{df['Durasi (Jam)'].mean():.1f} Jam" if len(df)>0 else "0 Jam", "Evaluasi Kecepatan", "off")
        st.markdown("<br>", unsafe_allow_html=True)

    if df.empty:
        st.info("ℹ️ Tidak ada data untuk periode yang dipilih.")
        return

    # ==========================================
    # KARTU KPI YANG DIKEMBALIKAN FULL (10 KARTU)
    # ==========================================
    st.markdown("#### 🎯 Key Performance Indicators")
    tot_bm = len(df[df['Kabupaten'] == "Bolaang Mongondow"]); tot_bms = len(df[df['Kabupaten'] == "Bolaang Mongondow Selatan"])
    tot_sma = len(df[df['Target'].str.contains('PT SMA', na=False)]); tot_jrbm = len(df[df['Target'].str.contains('PT JRBM', na=False)])
    tot_lost_time = df['Durasi (Jam)'].sum(); tot_semua = len(df)
    tot_cepat = len(df[df['Kategori Durasi'] == "Cepat"]); tot_lambat = len(df[df['Kategori Durasi'] == "Lambat"])
    avg_bm = df[df['Kabupaten'] == "Bolaang Mongondow"]['Durasi (Jam)'].mean(); avg_bms = df[df['Kabupaten'] == "Bolaang Mongondow Selatan"]['Durasi (Jam)'].mean()

    with st.container():
        # Baris 1
        c1, c2, c3, c4 = st.columns(4)
        with c1: st.markdown(create_kpi_card("Total Blokade BM", tot_bm, "Kejadian Area BM", "card-bm"), unsafe_allow_html=True)
        with c2: st.markdown(create_kpi_card("Total Blokade BMS", tot_bms, "Kejadian Area BMS", "card-bms"), unsafe_allow_html=True)
        with c3: st.markdown(create_kpi_card("Blokade PT SMA", tot_sma, "Melibatkan PT SMA", "card-sma"), unsafe_allow_html=True)
        with c4: st.markdown(create_kpi_card("Blokade PT JRBM", tot_jrbm, "Melibatkan PT JRBM", "card-jrbm"), unsafe_allow_html=True)

        # Baris 2
        c5, c6, c7, c8 = st.columns(4)
        with c5: st.markdown(create_kpi_card("Total Lost Time", f"{tot_lost_time:.1f} Jam", "Waktu Terbuang", "card-lost"), unsafe_allow_html=True)
        with c6: st.markdown(create_kpi_card("Total Semua Blokade", tot_semua, "Keseluruhan Area", "card-total"), unsafe_allow_html=True)
        with c7: st.markdown(create_kpi_card("Penanganan Cepat", tot_cepat, "≤ 9 Jam", "card-cepat"), unsafe_allow_html=True)
        with c8: st.markdown(create_kpi_card("Penanganan Lambat", tot_lambat, "> 9 Jam", "card-lambat"), unsafe_allow_html=True)

        # Baris 3
        c9, c10, c11, c12 = st.columns(4)
        with c9: st.markdown(create_kpi_card("Rata-rata Durasi BM", f"{avg_bm:.1f} Jam" if pd.notna(avg_bm) else "0 Jam", "Rata-rata Waktu", "card-avg-bm"), unsafe_allow_html=True)
        with c10: st.markdown(create_kpi_card("Rata-rata Durasi BMS", f"{avg_bms:.1f} Jam" if pd.notna(avg_bms) else "0 Jam", "Rata-rata Waktu", "card-avg-bms"), unsafe_allow_html=True)

    st.markdown("---")
    
    # PETA INTERAKTIF
    st.markdown("#### 🗺️ Pemetaan Geospasial Titik Blokade")
    df_map = df.groupby(['Kabupaten', 'Desa']).size().reset_index(name='Jumlah Kasus')
    df_map['lat'] = df_map.apply(lambda r: get_coordinates(r['Desa'], r['Kabupaten'])[0], axis=1)
    df_map['lon'] = df_map.apply(lambda r: get_coordinates(r['Desa'], r['Kabupaten'])[1], axis=1)
    
    fig_map = px.scatter_mapbox(
        df_map, lat="lat", lon="lon", size="Jumlah Kasus", color="Kabupaten", 
        hover_name="Desa", hover_data=["Jumlah Kasus"],
        color_discrete_map={"Bolaang Mongondow": "#3b82f6", "Bolaang Mongondow Selatan": "#ef4444"},
        zoom=8, mapbox_style="carto-positron", title="Titik Persebaran Blokade"
    )
    fig_map.update_layout(margin={"r":0,"t":40,"l":0,"b":0}, paper_bgcolor="#ffffff")
    st.plotly_chart(fig_map, use_container_width=True)

    st.markdown("---")

    # ==========================================
    # GRAFIK ANALITIK ADVANCED (HEATMAP & TREND ISU)
    # ==========================================
    st.markdown("#### 🔥 Analisis Waktu Rawan & Tren Isu (Advanced)")
    
    col_adv1, col_adv2 = st.columns(2)
    with col_adv1:
        # Pre-processing untuk Heatmap
        df['Hari'] = pd.to_datetime(df['Tanggal']).dt.day_name()
        hari_map = {'Monday': 'Senin', 'Tuesday': 'Selasa', 'Wednesday': 'Rabu', 'Thursday': 'Kamis', 'Friday': 'Jumat', 'Saturday': 'Sabtu', 'Sunday': 'Minggu'}
        df['Hari'] = df['Hari'].map(hari_map)
        # Menarik jam dari Waktu Mulai (misal '08:00' jadi 8)
        df['Jam'] = pd.to_datetime(df['Waktu Mulai'], format='%H:%M:%S', errors='coerce').dt.hour
        
        if not df['Jam'].isna().all():
            heatmap_data = df.groupby(['Hari', 'Jam']).size().reset_index(name='Jumlah')
            # Custom sorting hari
            cats = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
            heatmap_data['Hari'] = pd.Categorical(heatmap_data['Hari'], categories=cats, ordered=True)
            heatmap_data = heatmap_data.sort_values('Hari')
            
            fig_heat = px.density_heatmap(heatmap_data, x="Jam", y="Hari", z="Jumlah", 
                                        color_continuous_scale="Reds", title="Heatmap Waktu Kejadian Rawan",
                                        labels={'Jam': 'Jam Kejadian (0-23)', 'Hari': 'Hari Kejadian'})
            st.plotly_chart(force_black_text_on_plot(fig_heat), use_container_width=True, theme=None)
        else:
            st.info("Format waktu tidak valid untuk membuat heatmap.")

    with col_adv2:
        if not df.empty:
            df_trend_isu = df.groupby(['Bulan_Tahun', 'Isu']).size().reset_index(name='Jumlah')
            # Konversi agar sorting di grafik sesuai urutan waktu beneran (tidak urut abjad)
            df_trend_isu['SortDate'] = pd.to_datetime(df_trend_isu['Bulan_Tahun'], format='%b %Y')
            df_trend_isu = df_trend_isu.sort_values('SortDate')
            
            fig_trend_isu = px.bar(df_trend_isu, x="Bulan_Tahun", y="Jumlah", color="Isu",
                                 title="Tren Pergerakan Isu per Bulan", text_auto=True)
            st.plotly_chart(force_black_text_on_plot(fig_trend_isu), use_container_width=True, theme=None)


    # GRAFIK ANALITIK DASAR
    st.markdown("#### 📊 Analitik Grafik Lanjutan")
    
    col_chart1, col_chart2 = st.columns(2)
    with col_chart1:
        isu_count = df['Isu'].value_counts().reset_index().head(3)
        isu_count.columns = ['Jenis Isu', 'Jumlah']
        fig_isu = px.bar(isu_count, x='Jenis Isu', y='Jumlah', title="Top 3 Jenis Isu Utama", text_auto=True, color='Jenis Isu', color_discrete_sequence=px.colors.qualitative.Set2)
        st.plotly_chart(force_black_text_on_plot(fig_isu), use_container_width=True, theme=None)

    with col_chart2:
        target_count = df['Target'].value_counts().reset_index()
        target_count.columns = ['Target Perusahaan', 'Jumlah']
        fig_target = px.pie(target_count, values='Jumlah', names='Target Perusahaan', title="Analisis Target Perusahaan", hole=0.45, color_discrete_sequence=px.colors.qualitative.Pastel)
        st.plotly_chart(force_black_text_on_plot(fig_target), use_container_width=True, theme=None)

    col_chart3, col_chart4 = st.columns(2)
    with col_chart3:
        df_bm = df[df['Kabupaten'] == "Bolaang Mongondow"]
        if not df_bm.empty:
            bm_desa = df_bm['Desa'].value_counts().reset_index()
            bm_desa.columns = ['Desa', 'Jumlah']
            fig_bm_desa = px.bar(bm_desa, x='Desa', y='Jumlah', title="Blokade By BM (Berdasarkan Desa)", text_auto=True, color_discrete_sequence=['#3b82f6'])
            st.plotly_chart(force_black_text_on_plot(fig_bm_desa), use_container_width=True, theme=None)

    with col_chart4:
        if not df_bm.empty:
            bm_isu = df_bm['Isu'].value_counts().reset_index()
            bm_isu.columns = ['Isu', 'Jumlah']
            fig_bm_isu = px.bar(bm_isu, x='Isu', y='Jumlah', title="Blokade By BM (Berdasarkan Jenis Isu)", text_auto=True, color_discrete_sequence=['#3b82f6'])
            st.plotly_chart(force_black_text_on_plot(fig_bm_isu), use_container_width=True, theme=None)

    col_chart_bms1, col_chart_bms2 = st.columns(2)
    with col_chart_bms1:
        df_bms = df[df['Kabupaten'] == "Bolaang Mongondow Selatan"]
        if not df_bms.empty:
            bms_desa = df_bms['Desa'].value_counts().reset_index()
            bms_desa.columns = ['Desa', 'Jumlah']
            fig_bms_desa = px.bar(bms_desa, x='Desa', y='Jumlah', title="Blokade By BMS (Berdasarkan Desa)", text_auto=True, color_discrete_sequence=['#ef4444'])
            st.plotly_chart(force_black_text_on_plot(fig_bms_desa), use_container_width=True, theme=None)

    with col_chart_bms2:
        if not df_bms.empty:
            bms_isu = df_bms['Isu'].value_counts().reset_index()
            bms_isu.columns = ['Isu', 'Jumlah']
            fig_bms_isu = px.bar(bms_isu, x='Isu', y='Jumlah', title="Blokade By BMS (Berdasarkan Jenis Isu)", text_auto=True, color_discrete_sequence=['#ef4444'])
            st.plotly_chart(force_black_text_on_plot(fig_bms_isu), use_container_width=True, theme=None)

    st.markdown("#### 🏢 Keterlibatan Perusahaan & Tren Waktu")
    col_chart5, col_chart6 = st.columns(2)

    with col_chart5:
        df_keterlibatan = pd.DataFrame({'Kategori Perusahaan': ['Keterlibatan PT JRBM', 'Keterlibatan PT SMA'], 'Jumlah Insiden': [tot_jrbm, tot_sma]})
        fig_keterlibatan = px.bar(df_keterlibatan, x='Kategori Perusahaan', y='Jumlah Insiden', title="Total Keterlibatan Perusahaan", text_auto=True, color='Kategori Perusahaan', color_discrete_map={'Keterlibatan PT JRBM': '#10b981', 'Keterlibatan PT SMA': '#f59e0b'})
        st.plotly_chart(force_black_text_on_plot(fig_keterlibatan), use_container_width=True, theme=None)

    with col_chart6:
        if not df.empty:
            df_time = df.sort_values('Tanggal').groupby('Tanggal')['Durasi (Jam)'].sum().reset_index()
            fig_time = px.line(df_time, x='Tanggal', y='Durasi (Jam)', markers=True, title="Pergerakan Total Lost Time (Jam)")
            fig_time.update_traces(line_color='#ef4444', marker=dict(size=8, color='#dc2828'))
            st.plotly_chart(force_black_text_on_plot(fig_time), use_container_width=True, theme=None)

    # WORD CLOUD (ANALISIS TEKS LOKASI & DESKRIPSI)
    if WORDCLOUD_AVAILABLE:
        st.markdown("#### ☁️ Analisis Teks (Word Cloud)")
        st.caption("Visualisasi kata-kata kunci yang sering muncul di lapangan berdasarkan input Deskripsi.")
        text_data = " ".join(df['Deskripsi'].dropna().astype(str))
        if text_data.strip():
            wordcloud = WordCloud(width=800, height=300, background_color='white', colormap='viridis', max_words=100).generate(text_data)
            fig_wc, ax_wc = plt.subplots(figsize=(10, 4))
            ax_wc.imshow(wordcloud, interpolation='bilinear')
            ax_wc.axis("off")
            st.pyplot(fig_wc)
        else:
            st.info("Belum ada data deskripsi yang cukup untuk Word Cloud.")

    # TABEL DETAIL & EXPORT MULTIFORMAT
    st.markdown("---")
    st.markdown("### 🗃️ Data Detail Blokade & Galeri")
    
    # Drop kolom bantuan perhitungan agar tabel tetap bersih
    display_df = df.drop(columns=['Tahun', 'Bulan_Tahun', 'Quarter', 'Hari', 'Jam', 'SortDate'], errors='ignore')
    if not display_df.empty: display_df['Tanggal'] = pd.to_datetime(display_df['Tanggal']).dt.strftime('%d-%m-%Y')
    
    col_d1, col_d2, col_d3 = st.columns([1, 1, 4])
    with col_d1:
        if FPDF_AVAILABLE:
            st.download_button("📄 Export PDF", data=generate_pdf(df), file_name="Laporan_Blokade.pdf", mime="application/pdf", type="primary")
    with col_d2:
        if EXCEL_AVAILABLE:
            st.download_button("📥 Export Excel", data=generate_excel(display_df), file_name="Laporan_Blokade.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
    
    st.dataframe(
        display_df, 
        use_container_width=True, 
        hide_index=True,
        column_config={
            "File_Bukti": st.column_config.TextColumn("File Bukti", help="Nama File Bukti yang tersimpan di server lokal", width="medium")
        }
    )

# ==========================================
# NAVIGASI UTAMA
# ==========================================
def main():
    inject_custom_css()
    
    if not st.session_state['logged_in']:
        login_page()
    else:
        # ==========================================================
        # Menambahkan Logo di Sidebar agar muncul di semua navigasi
        # ==========================================================
        try:
            st.sidebar.image(LOGO_PATH, use_container_width=True)
        except Exception:
            pass # Mengabaikan error jika gambar tidak ditemukan
        
        st.sidebar.markdown("<hr style='margin-top: 5px; margin-bottom: 15px;'>", unsafe_allow_html=True)

        st.sidebar.markdown(f"""
        <div class="sidebar-profile">
            <div style="font-size: 50px; margin-bottom: 5px;">🧑‍💼</div>
            <h4 style="margin: 0;">{st.session_state['nama_lengkap']}</h4>
            <p style="margin: 5px 0 0 0; font-size: 12px; color: #94a3b8 !important; border-top: 1px solid #334155; padding-top: 8px;">🟢 {st.session_state['role']}</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.sidebar.markdown("### 🧭 Navigasi Utama")
        
        if st.session_state['role'] == "Administrator":
            menu = ["📝 Input Data Blokade", "⚙️ Kelola Data (Admin)", "📊 Dashboard Analitik"]
        elif st.session_state['role'] == "Supervisor": 
            menu = ["📝 Input Data Blokade", "⚙️ Kelola Data (Admin)"]
        else: # Manager
            menu = ["📊 Dashboard Analitik", "⚙️ Kelola Data (Admin)"]

        choice = st.sidebar.radio("Pilih Menu:", menu, label_visibility="collapsed")
        
        st.sidebar.markdown("---")
        if st.sidebar.button("🚪 Logout / Keluar", use_container_width=True):
            add_audit_log(st.session_state['nama_lengkap'], "LOGOUT", "User telah keluar dari sistem.")
            st.session_state['logged_in'] = False
            st.session_state['role'] = None
            st.session_state['nama_lengkap'] = None
            st.rerun()

        if choice == "📝 Input Data Blokade": input_form_page()
        elif choice == "📊 Dashboard Analitik": dashboard_page()
        elif choice == "⚙️ Kelola Data (Admin)": kelola_data_page()

if __name__ == "__main__":
    main()
